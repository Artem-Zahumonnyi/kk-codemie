import requests
from openpyxl import Workbook

# Keycloak server details
keycloak_server_url = '<KEYCLOAK_URL>'
master_realm_name = '<MASTER_REALM>'
target_realm_name = 'CODEMIE_REALM'
client_id = 'admin-cli'  # or your custom client ID if you've set up a specific client for API access
username = '<USERNAME>'
password = '<PASSWORD>'

# URLs for authentication
token_url = f"{keycloak_server_url}realms/{master_realm_name}/protocol/openid-connect/token"

# Obtain an access token from the master realm
token_response = requests.post(token_url, data={
    'client_id': client_id,
    'username': username,
    'password': password,
    'grant_type': 'password'
})
token_response.raise_for_status()
access_token = token_response.json().get('access_token')

# Initialize workbook and sheet for output
wb = Workbook()
ws = wb.active
ws.title = "Users"
ws.append(['Name Email', 'Email', 'Last Name', 'First Name', 'Projects'])

all_projects = set()
users_fetched = 0
max_results_per_request = 100  # Adjust as needed, considering server performance and rate limits

users_data = []  # List to hold user data for sorting

while True:
    # Paginated fetching of users
    users_url = f"{keycloak_server_url}admin/realms/{target_realm_name}/users?first={users_fetched}&max={max_results_per_request}"
    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
    users_response = requests.get(users_url, headers=headers)
    users_response.raise_for_status()
    users = users_response.json()

    if not users:
        # Break the loop if no more users are returned
        break

    for user in users:
        email = user.get('email', '')
        last_name = user.get('lastName', '')
        first_name = user.get('firstName', '')

        # Initialize list to hold projects for the current user
        user_projects = []

        # Process only the 'applications' attribute
        applications = user.get('attributes', {}).get('applications', [])
        for val in applications:
            projects = [project.strip().upper() for project in val.split(',')]
            # Filter projects, excluding those ending with @epam.com
            filtered_projects = [project for project in projects if not project.endswith('@EPAM.COM')]
            user_projects.extend(filtered_projects)
            all_projects.update(filtered_projects)

        user_projects_str = ", ".join(sorted(user_projects))  # Sort projects for consistent ordering
        users_data.append({'email': email, 'last_name': last_name, 'first_name': first_name, 'projects_str': user_projects_str})

    users_fetched += len(users)

# Sort users_data by 'projects_str'
sorted_users_data = sorted(users_data, key=lambda x: x['projects_str'])

# Write sorted data to the worksheet
for user in sorted_users_data:
    ws.append([user['email'], user['email'], user['last_name'], user['first_name'], user['projects_str']])

# Add all unique projects to a new sheet
ws_projects = wb.create_sheet(title="Projects")
for project in sorted(all_projects):
    ws_projects.append([project])

# Save the workbook
wb.save("keycloak_users_codemie-prod.xlsx")

print("Script completed. Users and projects have been saved to 'keycloak_users_codemie-prod.xlsx'.")

#############################
# # URLs for authentication
# token_url = f"{keycloak_server_url}realms/{master_realm_name}/protocol/openid-connect/token"
#
# # Obtain an access token from the master realm
# token_response = requests.post(token_url, data={
#     'client_id': client_id,
#     'username': username,
#     'password': password,
#     'grant_type': 'password'
# })
# token_response.raise_for_status()
# access_token = token_response.json().get('access_token')
#
# # Initialize workbook and sheet for output
# wb = Workbook()
# ws = wb.active
# ws.title = "Users"
# ws.append(['Name Email', 'Email', 'Last Name', 'First Name', 'Projects'])
#
# all_projects = set()
# users_fetched = 0
# max_results_per_request = 100  # Adjust as needed, considering server performance and rate limits
#
# while True:
#     # Paginated fetching of users
#     users_url = f"{keycloak_server_url}admin/realms/{target_realm_name}/users?first={users_fetched}&max={max_results_per_request}"
#     headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
#     users_response = requests.get(users_url, headers=headers)
#     users_response.raise_for_status()
#     users = users_response.json()
#
#     if not users:
#         # Break the loop if no more users are returned
#         break
#
#     for user in users:
#         email = user.get('email', '')
#         last_name = user.get('lastName', '')
#         first_name = user.get('firstName', '')
#
#         # Initialize list to hold projects for the current user
#         user_projects = []
#
#         # Assuming the user's attributes contain application names with projects as values
#         for attr_key, attr_value in user.get('attributes', {}).items():
#             for val in attr_value:
#                 projects = [project.strip().upper() for project in val.split(',')]
#                 user_projects.extend(projects)
#                 all_projects.update(projects)
#
#         user_projects_str = ", ".join(user_projects)
#         ws.append([email, email, last_name, first_name, user_projects_str])
#
#     users_fetched += len(users)
#
# # Add all unique projects to a new sheet
# ws_projects = wb.create_sheet(title="Projects")
# for project in sorted(all_projects):
#     ws_projects.append([project])
#
# # Save the workbook
# wb.save("keycloak_users_codemie-prod.xlsx")
#
# print("Script completed. Users and projects have been saved to 'keycloak_users_codemie-prod.xlsx'.")
